from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook
import time
# url = "https://workey.codeit.kr/costagram/index"
# driver = webdriver.Chrome()
# driver.implicitly_wait(3)
# driver.get(url)
# #driver.find_element_by_css_selector() #soup.select_one과 비슷 버전4부터 사라짐
# driver.find_element(By.CSS_SELECTOR, ".top-nav__login-link").click()
# driver.find_element(By.CSS_SELECTOR, ".login-container__login-input").send_keys('codeit')
# driver.find_element(By.CSS_SELECTOR, ".login-container__password-input").send_keys('datascience')
# driver.find_element(By.CSS_SELECTOR, ".login-container__login-button").click()
# time.sleep(3)

# 연습 문제
# url = "https://workey.codeit.kr/music"
# driver = webdriver.Chrome()
# driver.implicitly_wait(3)
# driver.get(url)
# time.sleep(3)
# popular_artists = []
# for artist in driver.find_elements(By.CSS_SELECTOR, 'ul.popular__order li'):
#     popular_artists.append(artist.text.strip())
# print(popular_artists)

# 바꿔보기
# driver = webdriver.Chrome()
# driver.implicitly_wait(3)
# driver.get("https://workey.codeit.kr/orangebottle/index")
# branch_infos = []
# for branch in driver.find_elements(By.CSS_SELECTOR, 'div.branch'):
#     branch_name = branch.find_element(By.CSS_SELECTOR,'p.city').text.strip()
#     address = branch.find_element(By.CSS_SELECTOR,'p.address').text.strip()
#     phone_number = branch.find_element(By.CSS_SELECTOR,'span.phoneNum').text.strip()
#     branch_infos.append([branch_name,address,phone_number])
# print(branch_infos)

# wb= Workbook(write_only=True)
# ws = wb.create_sheet('플레이 리스트')
# ws.append(['제목','해시태그','좋아요 수','곡 수'])
# driver = webdriver.Chrome()
# driver.implicitly_wait(3)
# driver.get("https://workey.codeit.kr/music")
# time.sleep(3)
# #현재 scrollHeight 가져오기
# last_height = driver.execute_script("return document.body.scrollHeight")
# while True:
#     #scrollHeight까지 스크롤
#     driver.execute_script("window.scrollTo(0,document.body.scrollHeight);")
#     #새로운 내용 로딩될때까지 기다림
#     time.sleep(0.5)
#     #새로운 내용 로딩됐는지 확인
#     new_height = driver.execute_script("return document.body.scrollHeight")
#     if new_height == last_height:
#         break;
#     last_height = new_height
# playlists = driver.find_elements(By.CSS_SELECTOR,'div.playlist__meta')
# for playlist in playlists:
#     title = playlist.find_element(By.CSS_SELECTOR,'h3.title').text
#     hashtags = playlist.find_element(By.CSS_SELECTOR,'p.tags').text
#     like_count = playlist.find_element(By.CSS_SELECTOR,'span.data__like-count').text
#     music_count = playlist.find_element(By.CSS_SELECTOR,'span.data__music-count').text
#     ws.append([title,hashtags,like_count,music_count])
# driver.quit()
# wb.save("플레이리스트_정보.xlsx")

# 워크북 생성
wb = Workbook(write_only=True)
ws = wb.create_sheet()
ws.append(['이미지 주소', '내용', '해시태그', '좋아요 수', '댓글 수'])

# 웹 드라이버 설정
driver = webdriver.Chrome()
driver.implicitly_wait(3)

driver.get('https://workey.codeit.kr/costagram/index')
time.sleep(1)

# 로그인
driver.find_element(By.CSS_SELECTOR, '.top-nav__login-link').click()
time.sleep(1)

driver.find_element(By.CSS_SELECTOR, '.login-container__login-input').send_keys('codeit')
driver.find_element(By.CSS_SELECTOR, '.login-container__password-input').send_keys('datascience')

driver.find_element(By.CSS_SELECTOR, '.login-container__login-button').click()
time.sleep(1)

# 페이지 끝까지 스크롤
last_height = driver.execute_script("return document.body.scrollHeight")

while True:
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(0.5)

    new_height = driver.execute_script("return document.body.scrollHeight")
    if new_height == last_height:
        break
    last_height = new_height

# 모든 썸네일 요소 가져오기
posts = driver.find_elements(By.CSS_SELECTOR, '.post-list__post')

for post in posts:
    # 썸네일 클릭
    post.click()
    time.sleep(0.5)

    # 포스팅 정보 가져오기
    style_attr = driver.find_element(By.CSS_SELECTOR, '.post-container__image').get_attribute('style')
    image_url = style_attr.split('"')[1]
    content = driver.find_element(By.CSS_SELECTOR, '.content__text').text.strip()
    hashtags = driver.find_element(By.CSS_SELECTOR, '.content__tag-cover').text.strip()
    like_count = driver.find_element(By.CSS_SELECTOR, '.content__like-count').text.strip()
    comment_count = driver.find_element(By.CSS_SELECTOR, '.content__comment-count').text.strip()
    ws.append([image_url, content, hashtags, like_count, comment_count])

    # 닫기 버튼 클릭
    driver.find_element(By.CSS_SELECTOR, '.close-btn').click()
    time.sleep(0.5)

driver.quit()

wb.save('코스타그램.xlsx')
